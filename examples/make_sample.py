from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Border, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

out = Path(__file__).with_name("jtc_screen_design.xlsx")
wb = Workbook()
ws = wb.active
ws.title = "画面設計書"
ws.merge_cells("B2:H2")
ws["B2"] = "画面設計書：ログイン画面"
ws["B2"].fill = PatternFill("solid", fgColor="D9EAF7")
for offset, value in enumerate(["項目", "内容", "必須", "入力方式", "備考"], start=2):
    ws.cell(4, offset).value = value
for row_index, row in enumerate([
    ["ユーザーID", "社員番号またはメール", "○", "テキスト", "半角英数"],
    ["パスワード", "8文字以上", "○", "パスワード", "マスク表示"],
    ["ログイン保持", "次回から自動ログイン", "", "チェックボックス", "任意"],
], start=5):
    for col_index, value in enumerate(row, start=2):
        ws.cell(row_index, col_index).value = value
thin = Side(style="thin", color="000000")
for row in ws.iter_rows(min_row=4, max_row=7, min_col=2, max_col=6):
    for cell in row:
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
dv = DataValidation(type="list", formula1='"テキスト,パスワード,チェックボックス,ラジオ"')
ws.add_data_validation(dv)
dv.add("E5:E7")
ws["F5"].comment = Comment("DB項目 user_id に対応", "architect")
wb.save(out)
print(out)
