import subprocess
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Border, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

from jtc_excel_md.converter import convert_workbook


def build_jtc_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "画面設計書"

    ws.merge_cells("B2:H2")
    ws["B2"] = "画面設計書：ログイン画面"
    ws["B2"].fill = PatternFill("solid", fgColor="D9EAF7")

    headers = ["項目", "内容", "必須", "入力方式", "備考"]
    for offset, value in enumerate(headers, start=2):
        ws.cell(4, offset).value = value
    rows = [
        ["ユーザーID", "社員番号またはメール", "○", "テキスト", "半角英数"],
        ["パスワード", "8文字以上", "○", "パスワード", "マスク表示"],
        ["ログイン保持", "次回から自動ログイン", "", "チェックボックス", "任意"],
    ]
    for row_index, row in enumerate(rows, start=5):
        for col_index, value in enumerate(row, start=2):
            ws.cell(row_index, col_index).value = value

    thin = Side(style="thin", color="000000")
    for row in ws.iter_rows(min_row=4, max_row=7, min_col=2, max_col=6):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col_index in range(2, 7):
        ws.cell(4, col_index).fill = PatternFill("solid", fgColor="EEEEEE")

    dv = DataValidation(type="list", formula1='"テキスト,パスワード,チェックボックス,ラジオ"')
    ws.add_data_validation(dv)
    dv.add("E5:E7")
    ws["F5"].comment = Comment("DB項目 user_id に対応", "architect")

    wb.save(path)


def test_convert_workbook_detects_merged_titles_bordered_tables_validations_and_comments(tmp_path: Path):
    source = tmp_path / "jtc.xlsx"
    build_jtc_workbook(source)

    result = convert_workbook(source)

    sheet = result["sheets"][0]
    assert sheet["name"] == "画面設計書"
    assert sheet["titles"] == [
        {
            "range": "B2:H2",
            "text": "画面設計書：ログイン画面",
            "start_cell": "B2",
        }
    ]

    table = sheet["blocks"][0]
    assert table["type"] == "bordered_table"
    assert table["range"] == "B4:F7"
    assert table["headers"] == ["項目", "内容", "必須", "入力方式", "備考"]
    assert table["rows"][1]["項目"] == "パスワード"
    assert table["rows"][1]["入力方式"] == "パスワード"
    assert table["cells"]["F5"]["comment"] == "DB項目 user_id に対応"

    assert sheet["validations"] == [
        {
            "range": "E5:E7",
            "type": "list",
            "options": ["テキスト", "パスワード", "チェックボックス", "ラジオ"],
        }
    ]


def test_convert_workbook_renders_specification_markdown_and_review_warnings(tmp_path: Path):
    source = tmp_path / "jtc.xlsx"
    build_jtc_workbook(source)

    result = convert_workbook(source)

    markdown = result["markdown"]
    assert "# 画面設計書：ログイン画面" in markdown
    assert "## 画面設計書 / B4:F7" in markdown
    assert "| 項目 | 内容 | 必須 | 入力方式 | 備考 |" in markdown
    assert "| パスワード | 8文字以上 | ○ | パスワード | マスク表示 |" in markdown
    assert "### 入力規則" in markdown
    assert "E5:E7: テキスト / パスワード / チェックボックス / ラジオ" in markdown

    warnings = result["warnings"]
    assert any("コメント付きセル F5" in warning for warning in warnings)
    assert any("結合セル B2:H2" in warning for warning in warnings)


def test_write_outputs_generates_preview_html_and_accuracy_report(tmp_path: Path):
    source = tmp_path / "jtc.xlsx"
    out_dir = tmp_path / "out"
    build_jtc_workbook(source)

    from jtc_excel_md.converter import write_outputs

    result = convert_workbook(source)
    write_outputs(result, out_dir)

    preview = (out_dir / "preview.html").read_text(encoding="utf-8")
    assert "JTC Excel Design Preview" in preview
    assert "画面設計書：ログイン画面" in preview
    assert "B4:F7" in preview
    assert 'data-cell="B5"' in preview
    assert "ユーザーID" in preview

    report = (out_dir / "evaluation.md").read_text(encoding="utf-8")
    assert "# Conversion Evaluation" in report
    assert "Sheets: 1" in report
    assert "Bordered blocks: 1" in report
    assert "Merged titles: 1" in report
    assert "Input validations: 1" in report
    assert "Warnings: 2" in report


def test_cli_module_writes_all_outputs(tmp_path: Path):
    source = tmp_path / "jtc.xlsx"
    out_dir = tmp_path / "out"
    build_jtc_workbook(source)

    completed = subprocess.run(
        [sys.executable, "-m", "jtc_excel_md.cli", str(source), "--out", str(out_dir)],
        check=True,
        capture_output=True,
        text=True,
    )

    for name in ["extracted.json", "specification.md", "warnings.md", "preview.html", "evaluation.md"]:
        assert (out_dir / name).exists()
        assert f"wrote: {out_dir / name}" in completed.stdout
