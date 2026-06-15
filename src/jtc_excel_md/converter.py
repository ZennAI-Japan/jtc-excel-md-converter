from __future__ import annotations

import json
import zipfile
from html import escape
from pathlib import Path
from typing import Any, cast

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

from .rich_media import extract_xlsx_drawing_media


def convert_workbook(path: str | Path) -> dict[str, Any]:
    """Convert a JTC-style Excel workbook into structured JSON plus Markdown."""
    workbook_path = Path(path)
    workbook = load_workbook(workbook_path, data_only=False)
    sheets = [_extract_sheet(sheet) for sheet in workbook.worksheets]
    rich_media = extract_xlsx_drawing_media(workbook_path)
    warnings = _collect_warnings(sheets, rich_media)
    markdown = _render_markdown(sheets, rich_media)
    return {
        "source": str(workbook_path),
        "sheets": sheets,
        "rich_media": rich_media,
        "markdown": markdown,
        "warnings": warnings,
    }


def write_outputs(result: dict[str, Any], output_dir: str | Path) -> None:
    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)
    (out / "extracted.json").write_text(
        json.dumps(_json_payload(result), ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    specification = result["markdown"] + "\n"
    (out / "book_specification.md").write_text(specification, encoding="utf-8")
    # Keep the original filename for backward compatibility with the initial CLI/README contract.
    (out / "specification.md").write_text(specification, encoding="utf-8")
    warnings = result.get("warnings") or ["要確認項目は検出されませんでした。"]
    (out / "warnings.md").write_text(
        "# Warnings\n\n" + "\n".join(f"- {warning}" for warning in warnings) + "\n",
        encoding="utf-8",
    )
    (out / "preview.html").write_text(_render_preview_html(result), encoding="utf-8")
    (out / "evaluation.md").write_text(_render_evaluation(result), encoding="utf-8")
    _write_package_zip(out)


def _write_package_zip(output_dir: Path) -> None:
    package_path = output_dir / "package.zip"
    artifact_names = [
        "book_specification.md",
        "specification.md",
        "extracted.json",
        "warnings.md",
        "preview.html",
        "evaluation.md",
    ]
    with zipfile.ZipFile(package_path, "w", compression=zipfile.ZIP_DEFLATED) as package:
        for name in artifact_names:
            path = output_dir / name
            if path.exists():
                package.write(path, arcname=name)


def _json_payload(result: dict[str, Any]) -> dict[str, Any]:
    return {key: value for key, value in result.items() if key != "markdown"}


def _extract_sheet(sheet: Worksheet) -> dict[str, Any]:
    titles = _extract_titles(sheet)
    blocks = _extract_bordered_tables(sheet)
    validations = _extract_validations(sheet)
    return {
        "name": sheet.title,
        "titles": titles,
        "blocks": blocks,
        "validations": validations,
    }


def _extract_titles(sheet: Worksheet) -> list[dict[str, str]]:
    titles: list[dict[str, str]] = []
    for merged in sheet.merged_cells.ranges:
        min_col, min_row, _max_col, _max_row = range_boundaries(str(merged))
        start_cell = f"{get_column_letter(min_col)}{min_row}"
        value = sheet.cell(min_row, min_col).value
        if value is None or str(value).strip() == "":
            continue
        titles.append({"range": str(merged), "text": str(value).strip(), "start_cell": start_cell})
    return titles


def _extract_bordered_tables(sheet: Worksheet) -> list[dict[str, Any]]:
    bordered = {
        (cell.row, cell.column)
        for row in sheet.iter_rows()
        for cell in row
        if _has_border(cell) and _not_empty_region_candidate(cell)
    }
    components = _connected_components(bordered)
    tables: list[dict[str, Any]] = []
    for component in components:
        min_row = min(row for row, _col in component)
        max_row = max(row for row, _col in component)
        min_col = min(col for _row, col in component)
        max_col = max(col for _row, col in component)
        if (max_row - min_row + 1) < 2 or (max_col - min_col + 1) < 2:
            continue
        headers = [_cell_text(sheet.cell(min_row, col)) for col in range(min_col, max_col + 1)]
        if not any(headers):
            continue
        rows: list[dict[str, str]] = []
        cells: dict[str, dict[str, str]] = {}
        for row_index in range(min_row + 1, max_row + 1):
            values = [_cell_text(sheet.cell(row_index, col)) for col in range(min_col, max_col + 1)]
            if not any(values):
                continue
            rows.append({header or f"列{idx + 1}": value for idx, (header, value) in enumerate(zip(headers, values))})
        for row_index, col_index in sorted(component):
            cell = sheet.cell(row_index, col_index)
            coordinate = cell.coordinate
            cell_payload: dict[str, str] = {}
            text = _cell_text(cell)
            if text:
                cell_payload["value"] = text
            if cell.comment:
                cell_payload["comment"] = str(cell.comment.text).strip()
            if cell.fill and cell.fill.fill_type:
                color = getattr(cell.fill.fgColor, "rgb", None)
                if color:
                    cell_payload["fill"] = str(color)
            if cell_payload:
                cells[coordinate] = cell_payload
        range_name = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
        tables.append(
            {
                "type": "bordered_table",
                "range": range_name,
                "headers": headers,
                "rows": rows,
                "cells": cells,
            }
        )
    return sorted(tables, key=lambda block: range_boundaries(block["range"])[1:3])


def _has_border(cell: Cell) -> bool:
    border = cell.border
    return any(getattr(side, "style", None) for side in [border.left, border.right, border.top, border.bottom])


def _not_empty_region_candidate(cell: Cell) -> bool:
    # Empty bordered cells are often part of JTC layout. Keep them only when the row/column is not completely blank.
    return True


def _connected_components(cells: set[tuple[int, int]]) -> list[set[tuple[int, int]]]:
    remaining = set(cells)
    components: list[set[tuple[int, int]]] = []
    while remaining:
        start = remaining.pop()
        stack = [start]
        component = {start}
        while stack:
            row, col = stack.pop()
            for neighbor in [(row - 1, col), (row + 1, col), (row, col - 1), (row, col + 1)]:
                if neighbor in remaining:
                    remaining.remove(neighbor)
                    component.add(neighbor)
                    stack.append(neighbor)
        components.append(component)
    return components


def _extract_validations(sheet: Worksheet) -> list[dict[str, Any]]:
    validations: list[dict[str, Any]] = []
    for validation in sheet.data_validations.dataValidation:
        options = _parse_validation_options(validation.formula1)
        for sqref in str(validation.sqref).split():
            validations.append({"range": sqref, "type": validation.type, "options": options})
    return validations


def _parse_validation_options(formula: str | None) -> list[str]:
    if not formula:
        return []
    stripped = str(formula).strip()
    if stripped.startswith('"') and stripped.endswith('"'):
        stripped = stripped[1:-1]
    if not stripped or stripped.startswith("="):
        return [stripped] if stripped else []
    return [part.strip() for part in stripped.split(",") if part.strip()]


def _cell_text(cell: Cell) -> str:
    value = cell.value
    if value is None:
        return ""
    text = str(value).strip()
    return text


def _collect_warnings(sheets: list[dict[str, Any]], rich_media: dict[str, list[dict[str, str]]] | None = None) -> list[str]:
    warnings: list[str] = []
    for sheet in sheets:
        for title in sheet["titles"]:
            warnings.append(f"{sheet['name']}: 結合セル {title['range']} を見出しとして解釈しました。")
        for block in sheet["blocks"]:
            for coordinate, payload in block["cells"].items():
                if "comment" in payload:
                    warnings.append(f"{sheet['name']}: コメント付きセル {coordinate} は人手確認してください。")
        for validation in sheet["validations"]:
            if validation["type"] == "list" and not validation["options"]:
                warnings.append(f"{sheet['name']}: 入力規則 {validation['range']} は外部範囲参照の可能性があります。")
    if rich_media:
        if rich_media.get("textboxes"):
            warnings.append("Excel内の図形/テキストボックスを抽出しました。セルとの位置関係は人手確認してください。")
        if rich_media.get("images"):
            warnings.append("Excel内の画像プレースホルダーを検出しました。画像内容は必要に応じて人手確認してください。")
        if rich_media.get("shapes"):
            warnings.append("Excel内の非テキスト図形プレースホルダーを検出しました。線・矢印・コネクタ等の意味は人手確認してください。")
    return warnings


def _render_markdown(sheets: list[dict[str, Any]], rich_media: dict[str, list[dict[str, str]]] | None = None) -> str:
    lines: list[str] = []
    for sheet in sheets:
        title = sheet["titles"][0]["text"] if sheet["titles"] else sheet["name"]
        lines.append(f"# {title}")
        lines.append("")
        if sheet["titles"]:
            lines.append("## 抽出した見出し")
            lines.append("")
            for title_info in sheet["titles"]:
                lines.append(f"- {title_info['range']}: {title_info['text']}")
            lines.append("")
        for block in sheet["blocks"]:
            lines.append(f"## {sheet['name']} / {block['range']}")
            lines.append("")
            headers = block["headers"]
            lines.append("| " + " | ".join(_escape_md(header or " ") for header in headers) + " |")
            lines.append("| " + " | ".join("---" for _ in headers) + " |")
            for row in block["rows"]:
                lines.append("| " + " | ".join(_escape_md(row.get(header or f"列{idx + 1}", "")) for idx, header in enumerate(headers)) + " |")
            lines.append("")
        if sheet["validations"]:
            lines.append("### 入力規則")
            lines.append("")
            for validation in sheet["validations"]:
                option_text = " / ".join(validation["options"]) if validation["options"] else "外部参照または未解析"
                lines.append(f"- {validation['range']}: {option_text}")
            lines.append("")
    if rich_media:
        if rich_media.get("textboxes"):
            lines.extend(["## 図形・テキストボックス", ""])
            for item in rich_media["textboxes"]:
                lines.append(f"- {item['drawing']} / {item['name']}: {item['text']}")
            lines.append("")
        if rich_media.get("images"):
            lines.extend(["## 画像プレースホルダー", ""])
            for item in rich_media["images"]:
                lines.append(f"- {item['drawing']} / {item['name']}")
            lines.append("")
        if rich_media.get("shapes"):
            lines.extend(["## 図形プレースホルダー", ""])
            for item in rich_media["shapes"]:
                lines.append(f"- {item['drawing']} / {item['name']}")
            lines.append("")
    return "\n".join(lines).strip()


def _escape_md(value: str) -> str:
    return value.replace("|", "\\|").replace("\n", "<br>")


def _render_preview_html(result: dict[str, Any]) -> str:
    if result.get("source_type") == "word_document":
        return _render_word_preview_html(result)
    sections: list[str] = []
    for sheet in result["sheets"]:
        sections.append(f"<section><h2>{escape(sheet['name'])}</h2>")
        for title in sheet["titles"]:
            sections.append(
                "<p class=\"title\">"
                f"<span>{escape(title['range'])}</span>{escape(title['text'])}"
                "</p>"
            )
        for block in sheet["blocks"]:
            sections.append(f"<h3>{escape(block['range'])}</h3>")
            sections.append("<table>")
            sections.append(
                "<thead><tr>"
                + "".join(f"<th>{escape(header)}</th>" for header in block["headers"])
                + "</tr></thead>"
            )
            sections.append("<tbody>")
            headers = block["headers"]
            cell_coordinates = _block_coordinates(block)
            for row_index, row in enumerate(block["rows"]):
                sections.append("<tr>")
                for col_index, header in enumerate(headers):
                    coordinate = cell_coordinates[row_index][col_index] if row_index < len(cell_coordinates) else ""
                    value = row.get(header or f"列{col_index + 1}", "")
                    sections.append(f'<td data-cell="{escape(coordinate)}">{escape(value)}</td>')
                sections.append("</tr>")
            sections.append("</tbody></table>")
        sections.append("</section>")
    body = "\n  ".join(sections)
    return f"""<!doctype html>
<html lang="ja">
<head>
  <meta charset="utf-8">
  <title>JTC Excel Design Preview</title>
  <style>
    body {{ font-family: -apple-system, BlinkMacSystemFont, "Hiragino Sans", sans-serif; margin: 32px; color: #172033; }}
    table {{ border-collapse: collapse; margin: 12px 0 28px; width: 100%; }}
    th, td {{ border: 1px solid #9aa4b2; padding: 8px 10px; vertical-align: top; }}
    th {{ background: #eef2f7; text-align: left; }}
    .title {{ background: #f4f7fb; border-left: 4px solid #315b8c; padding: 10px 12px; }}
    .title span {{ color: #667085; display: inline-block; margin-right: 12px; font-family: ui-monospace, monospace; }}
    [data-cell]::before {{ content: attr(data-cell); color: #98a2b3; font-size: 11px; display: block; }}
  </style>
</head>
<body>
  <h1>JTC Excel Design Preview</h1>
  <p>Extracted preview for human review. Cell coordinates are shown for traceability.</p>
  {body}
</body>
</html>
"""


def _render_word_preview_html(result: dict[str, Any]) -> str:
    document = result.get("document") or {}
    paragraphs = document.get("paragraphs", []) if isinstance(document, dict) else []
    tables = document.get("tables", []) if isinstance(document, dict) else []
    sections: list[str] = []
    if isinstance(paragraphs, list):
        sections.append("<section><h2>Paragraphs</h2>")
        for paragraph in paragraphs:
            if not isinstance(paragraph, dict):
                continue
            style = paragraph.get("style") or "normal"
            text = str(paragraph.get("text") or "")
            sections.append(
                f'<p class="paragraph"><span>{escape(str(style))}</span>{escape(text).replace(chr(10), "<br>")}</p>'
            )
        sections.append("</section>")
    if isinstance(tables, list):
        for table in tables:
            if not isinstance(table, dict):
                continue
            rows = table.get("rows") or []
            if not isinstance(rows, list):
                continue
            sections.append(f"<section><h2>Table {escape(str(table.get('index') or ''))}</h2><table>")
            for row in rows:
                if not isinstance(row, list):
                    continue
                sections.append("<tr>" + "".join(f"<td>{escape(str(cell))}</td>" for cell in row) + "</tr>")
            sections.append("</table></section>")
    body = "\n  ".join(sections)
    return f"""<!doctype html>
<html lang="ja">
<head>
  <meta charset="utf-8">
  <title>Word Document Preview</title>
  <style>
    body {{ font-family: -apple-system, BlinkMacSystemFont, "Hiragino Sans", sans-serif; margin: 32px; color: #172033; }}
    table {{ border-collapse: collapse; margin: 12px 0 28px; width: 100%; }}
    td {{ border: 1px solid #9aa4b2; padding: 8px 10px; vertical-align: top; }}
    .paragraph {{ background: #f8fafc; border-left: 4px solid #315b8c; padding: 10px 12px; }}
    .paragraph span {{ color: #667085; display: inline-block; margin-right: 12px; font-family: ui-monospace, monospace; }}
  </style>
</head>
<body>
  <h1>Word Document Preview</h1>
  <p>Extracted preview for human review. Paragraph styles and table text are shown for traceability.</p>
  {body}
</body>
</html>
"""


def _block_coordinates(block: dict[str, Any]) -> list[list[str]]:
    bounds = range_boundaries(block["range"])
    assert all(bound is not None for bound in bounds)
    min_col, min_row, max_col, max_row = cast(tuple[int, int, int, int], bounds)
    return [
        [f"{get_column_letter(col)}{row}" for col in range(min_col, max_col + 1)]
        for row in range(min_row + 1, max_row + 1)
    ]


def _render_evaluation(result: dict[str, Any]) -> str:
    sheets = result["sheets"]
    block_count = sum(len(sheet["blocks"]) for sheet in sheets)
    title_count = sum(len(sheet["titles"]) for sheet in sheets)
    validation_count = sum(len(sheet["validations"]) for sheet in sheets)
    warning_count = len(result.get("warnings", []))
    lines = [
            "# Conversion Evaluation",
            "",
            f"- Sheets: {len(sheets)}",
            f"- Bordered blocks: {block_count}",
            f"- Merged titles: {title_count}",
            f"- Input validations: {validation_count}",
            f"- Warnings: {warning_count}",
            "",
            "## Review Notes",
            "",
            "Use `warnings.md` to review ambiguous or manually confirmed items before using the Markdown as a downstream specification source.",
            "",
        ]
    if "ai" in result:
        ai = result["ai"]
        lines.extend(
            [
                "## AI Configuration",
                "",
                f"- Enabled: {ai.get('enabled')}",
                f"- Provider: {ai.get('provider')}",
                f"- Model: {ai.get('model')}",
                f"- Base URL: {ai.get('base_url')}",
                f"- API key configured: {ai.get('api_key_configured')}",
                "",
            ]
        )
    return "\n".join(lines)
